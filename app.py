import sqlite3
from flask import Flask, render_template, request, redirect, send_file
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import Workbook
import base64
import uuid
import os
import sqlite3
from openpyxl.drawing.image import Image as XLImage

def get_departments():
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "SELECT short_name, full_name FROM departments ORDER BY short_name"
    )
    rows = cur.fetchall()
    conn.close()
    return rows

def get_db():
    conn = sqlite3.connect("report.db")
    conn.row_factory = sqlite3.Row
    return conn
    
def get_department_full(code):
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "SELECT full_name FROM departments WHERE short_name = ?",
        (code,)
    )
    row = cur.fetchone()
    conn.close()
    return row[0] if row else code


app = Flask(__name__, static_folder="static", static_url_path="/static")
app.secret_key = "report_app_secret_key"
DB_NAME = "report.db"


# ==================================================
# สร้างเลขงาน
# ==================================================
def generate_work_no(created_at):
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()

    dt = datetime.strptime(created_at, "%Y-%m-%d %H:%M:%S")
    month = dt.strftime("%m")
    year = dt.strftime("%y")
    suffix = f"{month}{year}"

    # ✅ หาเลขล่าสุดของเดือนนี้
    cursor.execute("""
        SELECT work_no
        FROM reports
        WHERE work_no LIKE ?
        ORDER BY work_no DESC
        LIMIT 1
    """, (f"%{suffix}",))

    row = cursor.fetchone()
    conn.close()

    if row:
        last_seq = int(row[0][:3])
        next_seq = last_seq + 1
    else:
        next_seq = 1

    return f"{str(next_seq).zfill(3)}{suffix}"


# ==================================================
# แปลงวันที่
# ==================================================
def format_date_th(date_str):
    if not date_str:
        return ""
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d %H:%M")
        year_be = dt.year + 543
        return f"{dt.day:02d}/{dt.month:02d}/{year_be} {dt.strftime('%H:%M')}"
    except:
        return ""

def thai_month_year(month, year):
    thai_months = [
        "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน",
        "พฤษภาคม", "มิถุนายน", "กรกฎาคม", "สิงหาคม",
        "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"
    ]
    thai_year = year + 543
    return f"{thai_months[month-1]} {thai_year}"


def format_date_th_short(date_str):
    if not date_str:
        return ""
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        thai_months = [
            "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน",
            "พฤษภาคม", "มิถุนายน", "กรกฎาคม", "สิงหาคม",
            "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"
        ]
        return f"{dt.day} {thai_months[dt.month-1]} {dt.year + 543}"
    except:
        return date_str
    
def format_month_th(year, month):
    months = [
        "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน",
        "พฤษภาคม", "มิถุนายน", "กรกฎาคม", "สิงหาคม",
        "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"
    ]
    return f"{months[month-1]} {year + 543}"



# ==================================================
# DB
# ==================================================
def init_db():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            work_no TEXT UNIQUE,
            receive_datetime TEXT,
            department TEXT,
            reporter TEXT,
            job_type TEXT,
            asset_no TEXT,
            problem TEXT,
            solution TEXT,
            completed_datetime TEXT,
            close_note TEXT,
            confirm_name TEXT,
            signature TEXT,
            created_at TEXT
        )
    """)
    
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            staff_name TEXT NOT NULL,
            work_date TEXT NOT NULL,
            time_in TEXT,
            time_out TEXT,
            note TEXT
        )
    """)
    conn.commit()
    conn.close()


def init_assets_db():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS assets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            asset_no TEXT UNIQUE NOT NULL,
            asset_type TEXT NOT NULL,
            asset_model TEXT,
            serial_no TEXT,
            mac_address TEXT,
            hostname TEXT,
            owner_name TEXT,
            position TEXT,
            department TEXT,
            status TEXT DEFAULT 'ใช้งาน',
            note TEXT
        )
    """)
    conn.commit()
    conn.close()



# ==================================================
# FIX ข้อมูลลายเซ็นที่เคยสลับช่อง
# ==================================================
def fix_signature_column():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute(
        """
        UPDATE reports
        SET signature = confirm_name,
            confirm_name = ''
        WHERE confirm_name LIKE '%.png'
          AND (signature IS NULL OR signature = '')
    """
    )
    conn.commit()
    conn.close()

DEPT_MAP = {
        "กลุ่มกฏหมาย":"กกม.",
        "กลุ่มพัฒนาระบบบริหาร":"กพร.",
        "กองส่งเสริมสิทธิและเสรีภาพ":"กสส.",
        "กองยุทธศาสตร์และแผนสิทธิมนุษยชนแห่งชาติ":"กยผ.",
        "สำนักงานป้องกันและปราบปรามการทรมานและการกระทำให้บุคคลสูญหาย":"สปท.",
        "กลุ่มตรวจสอบภายใน":"กตน.",
        "กองสิทธิมนุษยชนระหว่างประเทศ":"กสป.",
        "สำนักงานช่วยเหลือทางการเงินแก่ผู้เสียหายและจำเลยในคดีอาญา":"สชง.",
        "กองส่งเสริมการระงับข้อพิพาท":"กสร.",
        "กองพิทักษ์สิทธิและเสรีภาพ":"กพส.",
        "สำนักงานคุ้มครองพยาน":"สคพ.",
        "กลุ่มผู้เชี่ยวชาญ":"กชช.",
        "ศูนย์ข้อมูลและพัฒนาระบบเทคโนโลยีสารสนเทศ":"ศพท.",
        "กลุ่มงานคลัง":"กงค.",
        "กลุ่มงานบริหารและพัฒนาทรัพยากรบุคคล":"กบค.",
        "กลุ่มงานช่วยอำนวยการและสื่อสารองค์กร":"กชอ.",
        "กลุ่มงานบริหารพัสดุ":"กบพ.",
        "ห้องประชุม": "ห้องประชุม",
        "ห้องอธิบดี": "ห้องอธิบดี",
        "ห้องรองอธิบดี": "ห้องรองอธิบดี",
}
from openpyxl import load_workbook
from flask import flash, url_for





@app.route("/assets/import", methods=["GET"])
def assets_import_page():
    return render_template("assets_import.html")

@app.route("/assets/import", methods=["POST"])
def assets_import():
    import pandas as pd

    file = request.files.get("file")
    if not file:
        return "ไม่พบไฟล์", 400

    # ✅ อ่านชื่อหน่วยงานจากแถวที่ 1 (A1)
    meta_df = pd.read_excel(file, header=None)
    dept_full = str(meta_df.iloc[0, 0]).strip()
    dept_short = DEPT_MAP.get(dept_full, dept_full)

    # ✅ อ่านข้อมูลจริง (หัวคอลัมน์อยู่แถวที่ 2)
    df = pd.read_excel(file, header=1)
    df = df.fillna("")   # 👈 บรรทัดนี้แหละ แก้ nan ทั้งหมด

    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()

    count = 0

    for _, row in df.iterrows():
        try:
            cursor.execute("""
                INSERT INTO assets (
                    asset_no,
                    asset_type,
                    asset_model,
                    serial_no,
                    mac_address,
                    hostname,
                    owner_name,
                    position,
                    department,
                    status
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                str(row["เลขครุภัณฑ์"]),
                str(row["ประเภทเครื่อง"]),
                str(row["ยี่ห้อ/รุ่น"]),
                str(row["Serial Number"]),
                str(row["Mac Address"]),
                str(row["ชื่อเครื่อง"]),
                str(row["ผู้ครอบครอง"]),
                str(row["ตำแหน่ง"]),
                dept_short,   # ✅ ใช้ค่าจากไฟล์จริง
                "ใช้งาน"
            ))
            count += 1
        except Exception as e:
            print("❌ ข้ามแถว:", e)

    conn.commit()
    conn.close()

    return f"อิมพอร์ตสำเร็จ {count} รายการ ({dept_short})"


# =========================
# ลงเวลา
# =========================

@app.route("/attendance", methods=["GET"])
def attendance_page():
    now = datetime.now()

    # ===== เดือนปัจจุบัน =====
    from calendar import monthrange
    year = now.year
    month = now.month

    month_start = f"{year}-{month:02d}-01"
    last_day = monthrange(year, month)[1]
    month_end = f"{year}-{month:02d}-{last_day}"

    selected_staff = request.args.get("staff_name", "")

    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    if selected_staff is not None and selected_staff != "":
        cur.execute("""
            SELECT staff_name, work_date, time_in, time_out, note
            FROM attendance
            WHERE staff_name = ?
            AND work_date BETWEEN ? AND ?
            ORDER BY work_date DESC
        """, (selected_staff, month_start, month_end))
    else:
        cur.execute("""
            SELECT staff_name, work_date, time_in, time_out, note
            FROM attendance
            WHERE work_date BETWEEN ? AND ?
            ORDER BY work_date DESC, staff_name
        """, (month_start, month_end))


    records = cur.fetchall()
    conn.close()

    today_th = format_date_th_short(now.strftime("%Y-%m-%d"))

    return render_template(
        "attendance.html",
        records=records,
        selected_staff=selected_staff,
        today_th=today_th,
        format_date_th_short=format_date_th_short
    )
    
from openpyxl import Workbook
from flask import send_file
import io
from calendar import monthrange

# =========================
# export เวลาทำงาน
# =========================
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.drawing.image import Image as XLImage
from calendar import monthrange
import io

@app.route("/attendance/export")
def export_attendance_excel():
    staff_name = request.args.get("staff_name")
    if not staff_name:
        return "กรุณาเลือกชื่อเจ้าหน้าที่", 400

    now = datetime.now()
    year = now.year
    month = now.month

    last_day = monthrange(year, month)[1]
    month_start = f"{year}-{month:02d}-01"
    month_end = f"{year}-{month:02d}-{last_day}"

    # ===== ดึงข้อมูล =====
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("""
        SELECT work_date, time_in, time_out
        FROM attendance
        WHERE staff_name = ?
          AND work_date BETWEEN ? AND ?
    """, (staff_name, month_start, month_end))
    rows = cur.fetchall()
    conn.close()

    attendance_map = {int(r["work_date"].split("-")[2]): r for r in rows}

    # ===== Workbook =====
    wb = Workbook()
    ws = wb.active
    ws.title = "ใบลงเวลา"

    # ===== ตั้งค่าหน้ากระดาษ A4 แนวตั้ง + บีบ 1 หน้า =====
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_setup.scale = None   # สำคัญ: ให้ fit ทำงานจริง

    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.6
    ws.page_margins.bottom = 0.6

    # ===== Style =====
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    bold = Font(bold=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ===== คอลัมน์ A = โลโก้ =====
    ws.column_dimensions["A"].width = 16

    logo = XLImage("static/img/logo.png")
    logo.height = 60
    logo.width = 115
    ws.add_image(logo, "A1")

    # ===== หัวกระดาษ =====
    ws.merge_cells("B1:F1")
    ws.merge_cells("B2:F2")
    ws.merge_cells("B3:F3")
    ws.merge_cells("C4:F4")

    ws["B1"] = "ใบลงเวลาปฏิบัติงานของเจ้าหน้าที่ประจำหน่วยงาน"
    ws["B1"].alignment = center
    ws["B1"].font = Font(bold=True, size=13)

    ws["B2"] = "กรมคุ้มครองสิทธิและเสรีภาพ"
    ws["B2"].alignment = center
    ws["B2"].font = Font(bold=True, size=12)

    ws["B3"] = f"ประจำเดือน {format_month_th(year, month)}"
    ws["B3"].alignment = center
    ws["B3"].font = Font(size=11)

    ws["C4"] = f"ชื่อเจ้าหน้าที่ : {staff_name}"
    ws["C4"].alignment = left
    ws["C4"].font = Font(size=11)

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20

    # ===== ตาราง =====
    start_row = 6
    ws.row_dimensions[start_row].height = 22

    headers = ["วันที่", "เวลาเข้า", "เวลาออก", "ลายมือชื่อ", "หมายเหตุ"]
    widths = {
        1: 6,
        2: 12,
        3: 12,
        4: 22,
        5: 22
    }

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.alignment = center
        cell.font = bold
        cell.border = border
        ws.column_dimensions[chr(64 + col)].width = widths[col]

    # ===== ข้อมูลรายวัน =====
    import os

    signature_path = f"static/signatures/{staff_name}.png"

    # กันพัง ถ้าไม่มีไฟล์ลายเซ็น
    if not os.path.exists(signature_path):
        signature_path = None

    for day in range(1, last_day + 1):
        r = start_row + day

        ws.cell(row=r, column=1, value=day).alignment = center

        record = attendance_map.get(day)
        if record:
            time_in = record["time_in"]
            time_out = record["time_out"]

            ws.cell(row=r, column=2, value=time_in or "")
            ws.cell(row=r, column=3, value=time_out or "")

            # ใส่ลายเซ็นเฉพาะวันที่มาทำงานจริง
            if time_in or time_out:
                if signature_path:
                    sig = XLImage(signature_path)
                    sig.width = 160 ##### 838 x 130 #####
                    sig.height = 16
                    ws.add_image(sig, f"D{r}")

        # เส้นตาราง (ทุกวัน)
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = border
            ws.cell(row=r, column=c).alignment = center





        # ===== ลายเซ็น =====
        sign_row = start_row + last_day + 2
        ws.merge_cells(f"B{sign_row}:F{sign_row}")
        ws.row_dimensions[sign_row].height = 24
        ws[f"B{sign_row}"] = "ลายเซ็นเจ้าหน้าที่ ............................................................."
        ws[f"B{sign_row}"].alignment = center
        ws[f"B{sign_row}"].font = Font(size=11)

        name_row = sign_row + 1

        ws.merge_cells(f"B{name_row}:F{name_row}")
        ws.row_dimensions[name_row].height = 22
        ws[f"B{name_row}"] = "(....................................................................................)"
        ws[f"B{name_row}"].alignment = center
        ws[f"B{name_row}"].font = Font(size=10)
            # ===== ส่งไฟล์ =====
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"ใบลงเวลา_{staff_name}_{format_month_th(year, month)}.xlsx"

    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )




from flask import session, request, redirect, render_template, url_for


# =========================
# หน้าเมนูกลาง (INDEX)
# =========================
@app.route("/")
def index():
    if not session.get("unlocked"):
        return redirect("/unlock")

    return render_template("index.html")
  
#==================================================
# ส่งข้อมูลไปที่หน้า report
#==================================================
@app.route("/report", endpoint="report_form")
def report_form():
    departments = get_departments()
    return render_template(
        "report.html",
        departments=departments
    )

# ==================================================
# เพิ่มงาน
# ==================================================
@app.route("/report", methods=["GET", "POST"], endpoint="save_report")
def save_report():
    if request.method == "POST":
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()

        created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        work_no = generate_work_no(created_at)

        receive_date = request.form.get("receive_date")
        receive_time = request.form.get("receive_time")

        receive_datetime = None
        if receive_date and receive_time:
            receive_datetime = f"{receive_date} {receive_time}"

        completed_datetime = None
        completed_date = request.form.get("complete_date")
        completed_time = request.form.get("complete_time")

        if completed_date and completed_time:
            completed_datetime = f"{completed_date} {completed_time}"

        # ===== ลายเซ็น =====
        signature_filename = None
        signature_data = request.form.get("signature")
        if signature_data and "," in signature_data:
            try:
                img_bytes = base64.b64decode(signature_data.split(",")[1])
                os.makedirs("static/signatures", exist_ok=True)
                signature_filename = f"{uuid.uuid4().hex}.png"
                with open(f"static/signatures/{signature_filename}", "wb") as f:
                    f.write(img_bytes)
            except:
                signature_filename = None

        confirm_name = request.form.get("confirm_name", "").strip()

        # ===== ดึงชื่อเต็มของหน่วยงาน =====
        department_short = request.form["department"]

        

        cursor.execute(
            """
            INSERT INTO reports (
                work_no,
                receive_datetime,
                department,
                reporter,
                job_type,
                asset_no,
                problem,
                solution,
                completed_datetime,
                close_note,
                confirm_name,
                signature,
                created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                work_no,
                receive_datetime,
                department_short,
                request.form["reporter"],
                request.form["job_type"],
                request.form.get("asset_no"),
                request.form.get("problem"),
                request.form.get("solution"),
                completed_datetime,
                request.form.get("close_note"),
                confirm_name,
                signature_filename,
                created_at,
            ),
        )

        conn.commit()
        fix_signature_column()
        conn.close()

        return redirect("/list?success=save")

    return render_template("report.html")



# ==================================================
# รายการ
# ==================================================
@app.route("/list")
def list_reports():
    from datetime import datetime
    import math

    PER_PAGE = 10

    page = request.args.get("page", 1, type=int)
    q = request.args.get("q", "").strip()
    job_type = request.args.get("job_type", "")
    staff = request.args.get("staff")          # 👈 รับค่าจาก work_compare
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")

    now = datetime.now()

    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()

    where = "WHERE 1=1"
    params = []

    # ===== ค้นหาจากคำค้น =====
    if q:
        for w in q.split():
            where += """
                AND (
                    work_no LIKE ?
                    OR department LIKE ?
                    OR reporter LIKE ?
                    OR job_type LIKE ?
                    OR asset_no LIKE ?
                    OR problem LIKE ?
                    OR solution LIKE ?
                )
            """
            kw = f"%{w}%"
            params.extend([kw] * 7)

    # ===== กรองประเภทงาน =====
    if job_type:
        where += " AND job_type = ?"
        params.append(job_type)

    # ===== กรองเจ้าหน้าที่ (มาจาก work_compare) =====
    if staff:
        where += " AND confirm_name = ?"
        params.append(staff)

    # ===== กรองช่วงวันที่ =====
    if date_from and date_to:
        where += " AND receive_datetime >= ?"
        params.append(f"{date_from} 00:00")

        where += " AND receive_datetime <= ?"
        params.append(f"{date_to} 23:59")

    else:
        where += " AND receive_datetime IS NOT NULL AND receive_datetime != ''"

        current_year = now.strftime("%Y")
        current_month = now.strftime("%m")

        where += " AND substr(receive_datetime, 1, 4) = ?"
        params.append(current_year)

        where += " AND substr(receive_datetime, 6, 2) = ?"
        params.append(current_month)

    # ===== COUNT =====
    cursor.execute(f"SELECT COUNT(*) FROM reports {where}", params)
    total_rows = cursor.fetchone()[0]
    total_pages = math.ceil(total_rows / PER_PAGE) if total_rows else 1

    if page > total_pages:
        page = total_pages
    if page < 1:
        page = 1

    offset = (page - 1) * PER_PAGE

    # ===== SELECT =====
    cursor.execute(f"""
        SELECT
            id,
            work_no,
            receive_datetime,
            department,
            job_type,
            problem,
            solution,
            reporter
        FROM reports
        {where}
        ORDER BY receive_datetime DESC
        LIMIT ? OFFSET ?
    """, params + [PER_PAGE, offset])

    raw = cursor.fetchall()
    conn.close()

    reports = raw

    return render_template(
        "list.html",
        reports=reports,
        get_department_full=get_department_full,
        format_date_th=format_date_th,
        page=page,
        total_pages=total_pages,
        keyword=q,
        job_type=job_type,
        date_from=date_from,
        date_to=date_to,
        staff=staff,   # 👈 ส่งให้ template ใช้ต่อ
    )


# ==================================================
# รายการสรุป
# ==================================================
@app.route("/report-summary", methods=["GET"])
def report_summary():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()

    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")

    is_search = False
    rows = []
    chart_data = {}

    total = software = hardware = network = other = 0
    date_from_th = date_to_th = ""

    if date_from and date_to:
        is_search = True

        date_from_th = format_date_th(date_from)
        date_to_th = format_date_th(date_to)

        # ===== ดึงข้อมูลตามช่วงวันที่ =====
        cursor.execute(
            """
            SELECT
                receive_datetime,
                department,
                job_type,
                problem,
                solution,
                reporter
            FROM reports
            WHERE date(receive_datetime)
                  BETWEEN date(?) AND date(?)
            ORDER BY receive_datetime ASC
        """,
            (date_from, date_to),
        )

        rows = cursor.fetchall()

        # ===== นับประเภทงาน =====
        for r in rows:
            total += 1
            if r[2] == "Software":
                software += 1
            elif r[2] == "Hardware":
                hardware += 1
            elif r[2] == "Network":
                network += 1
            else:
                other += 1

        # ===== เตรียมข้อมูลกราฟ =====
        dept_count = {}
        for r in rows:
            dept = r[1]
            dept_count.setdefault(
                dept, {"Software": 0, "Hardware": 0, "Network": 0, "Other": 0}
            )
            dept_count[dept][r[2]] += 1

        chart_data = {
            "labels": list(dept_count.keys()),
            "software": [v["Software"] for v in dept_count.values()],
            "hardware": [v["Hardware"] for v in dept_count.values()],
            "network": [v["Network"] for v in dept_count.values()],
            "other": [v["Other"] for v in dept_count.values()],
        }

    conn.close()

    return render_template(
        "report_summary.html",
        rows=rows,
        total=total,
        software=software,
        hardware=hardware,
        network=network,
        other=other,
        chart_data=chart_data,
        date_from=date_from,
        date_to=date_to,
        date_from_th=date_from_th,
        date_to_th=date_to_th,
        is_search=is_search,
        format_date_th=format_date_th,
    )
    
from collections import defaultdict
from datetime import datetime
import calendar
import sqlite3
from flask import request, render_template

@app.route("/report-monthly-summary")
def report_monthly_summary():

    # ===============================
    # 1) รับค่าเดือน / ปี
    # ===============================
    month = request.args.get("month", type=int)
    year = request.args.get("year", type=int)

    now = datetime.now()
    current_year = now.year

    if not month or not year:
        month = now.month
        year = now.year

    last_day = calendar.monthrange(year, month)[1]
    date_from = f"{year}-{month:02d}-01"
    date_to = f"{year}-{month:02d}-{last_day}"

    # ===============================
    # 2) ดึงข้อมูลจาก DB
    # ===============================
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("""
        SELECT
            department,
            job_type,
            COUNT(*) AS total
        FROM reports
        WHERE receive_datetime >= ?
          AND receive_datetime <= ?
        GROUP BY department, job_type
    """, (f"{date_from} 00:00", f"{date_to} 23:59"))

    rows = cur.fetchall()
    conn.close()

    # ===============================
    # 3) เตรียมโครงข้อมูล
    # ===============================
    types = ["Software", "Hardware", "Network", "Other"]

    summary = defaultdict(lambda: {
        "items": {t: 0 for t in types},
        "total": 0
    })

    grand_total = {
        "items": {t: 0 for t in types},
        "total": 0
    }

    # ===============================
    # 4) ใส่ข้อมูลลง summary
    # ===============================
    for r in rows:
        dept = (r["department"] or "ไม่ระบุ").strip()
        t = (r["job_type"] or "Other").strip().title()
        count = r["total"]

        if t not in types:
            t = "Other"

        summary[dept]["items"][t] += count
        summary[dept]["total"] += count

        grand_total["items"][t] += count
        grand_total["total"] += count

    # ===============================
    # 5) ส่งให้ template
    # ===============================
    return render_template(
        "report_monthly_summary.html",
        date_from=date_from,
        date_to=date_to,
        types=types,
        summary=summary,
        grand_total=grand_total,
        selected_month=month,
        selected_year=year,
        current_year=current_year
    )



   


# ==================================================
# รีพอร์ต excel
# ==================================================
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

TH_MONTHS = [
    "",
    "มกราคม",
    "กุมภาพันธ์",
    "มีนาคม",
    "เมษายน",
    "พฤษภาคม",
    "มิถุนายน",
    "กรกฎาคม",
    "สิงหาคม",
    "กันยายน",
    "ตุลาคม",
    "พฤศจิกายน",
    "ธันวาคม",
]


def parse_datetime_safe(date_str):
    """
    รองรับ:
    - YYYY-MM-DD HH:MM:SS
    - YYYY-MM-DD HH:MM
    - YYYY-MM-DD
    """
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(date_str, fmt)
        except:
            pass
    return None


def format_date_full_th(date_str):
    dt = parse_datetime_safe(date_str)
    if not dt:
        return "-"
    return f"{dt.day} {TH_MONTHS[dt.month]} {dt.year + 543}"


def format_time_th(date_str):
    dt = parse_datetime_safe(date_str)
    if not dt:
        return "-"
    return dt.strftime("%H.%M น.")


@app.route("/export-excel", methods=["GET"])
def export_excel():
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")

    if not date_from or not date_to:
        return "กรุณาเลือกช่วงวันที่ก่อน Export", 400

    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()

    cursor.execute(
        """
        SELECT
            receive_datetime,
            asset_no,
            department,
            reporter,
            job_type,
            problem,
            solution
        FROM reports
        WHERE date(receive_datetime)
              BETWEEN date(?) AND date(?)
        ORDER BY receive_datetime ASC
    """,
        (date_from, date_to),
    )

    rows = cursor.fetchall()
    conn.close()

    # ===== Excel =====
    wb = Workbook()
    ws = wb.active
    ws.title = "รายงานแจ้งปัญหา"

    headers = [
        "ลำดับ",
        "วันที่",
        "เวลา",
        "เลขครุภัณฑ์",
        "หน่วยงาน",
        "ผู้แจ้ง",
        "ประเภทงาน",
        "ปัญหา",
        "วิธีแก้ไข",
    ]

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E7EDF8")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Header
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Data
    for idx, r in enumerate(rows, start=1):
        ws.append(
            [
                idx,
                format_date_full_th(r[0]),
                format_time_th(r[0]),
                r[1] or "",
                r[2],
                r[3],
                r[4],
                r[5],
                r[6],
            ]
        )

    # Style ทุกช่อง
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center
            cell.border = border

    # Row height เท่ากัน
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 28

    # Auto column width
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        if col_letter in ["H", "I"]:
            ws.column_dimensions[col_letter].width = 40
        else:
            ws.column_dimensions[col_letter].width = max_length + 4

        filename = f"report_{date_from}_to_{date_to}.xlsx"
        os.makedirs("reports", exist_ok=True)
        filepath = os.path.join("reports", filename)

        wb.save(filepath)


    return send_file(filepath, as_attachment=True, download_name=filename)


# ==================================================
# ดูรายละเอียด (แก้หลัก)
# ==================================================
@app.route("/view/<int:report_id>")
def view_report(report_id):
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()

    cursor.execute(
        """
        SELECT
            id,
            work_no,
            receive_datetime,
            department,
            reporter,
            job_type,
            asset_no,
            problem,
            solution,
            completed_datetime,
            close_note,
            confirm_name,
            signature
        FROM reports
        WHERE id=?
    """,
        (report_id,),
    )

    r = cursor.fetchone()
    conn.close()

    if not r:
        return "ไม่พบข้อมูลงาน", 404

    data = {
        "id": r[0],
        "work_no": r[1],
        "receive_datetime": format_date_th(r[2]),
        "department": get_department_full(r[3]),
        "reporter": r[4],
        "job_type": r[5],
        "asset_no": r[6],
        "problem": r[7],
        "solution": r[8],
        "completed_datetime": format_date_th(r[9]),
        "close_note": r[10],
        "confirm_name": r[11],
        "signature": os.path.basename(r[12]) if r[12] else None,
    }

    return render_template("view.html", data=data)


# ==================================================
# แก้ไข
# ==================================================
@app.route("/edit/<int:report_id>", methods=["GET", "POST"])
def edit_report(report_id):
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()

    if request.method == "POST":
        receive_datetime = (
            f"{request.form['receive_date']} {request.form['receive_time']}"
        )

        completed_datetime = None
        completed_date = request.form.get("complete_date")
        completed_time = request.form.get("complete_time")

        if completed_date and completed_time:
            completed_datetime = f"{completed_date} {completed_time}"

        cursor.execute(
            """
            UPDATE reports SET
                receive_datetime=?,
                department=?,
                reporter=?,
                job_type=?,
                asset_no=?,
                problem=?,
                solution=?,
                completed_datetime=?,
                close_note=?,
                confirm_name=?
            WHERE id=?
        """,
            (
                receive_datetime,
                request.form["department"],
                request.form["reporter"],
                request.form["job_type"],
                request.form.get("asset_no"),
                request.form.get("problem"),
                request.form.get("solution"),
                completed_datetime,
                request.form.get("close_note"),
                request.form.get("confirm_name"),
                report_id,
            ),
        )

        conn.commit()
        conn.close()
        return redirect("/list?success=edit")

    cursor.execute(
        """
        SELECT
            id, work_no, receive_datetime,
            department, reporter, job_type,
            asset_no, problem, solution,
            completed_datetime, close_note, confirm_name
        FROM reports
        WHERE id=?
    """,
        (report_id,),
    )

    r = cursor.fetchone()
    conn.close()

    receive_date, receive_time = r[2].split(" ")
    complete_date, complete_time = ("", "")
    if r[9]:
        complete_date, complete_time = r[9].split(" ")

    data = {
        "id": r[0],
        "work_no": r[1],
        "receive_date": receive_date,
        "receive_time": receive_time,
        "department": r[3],
        "reporter": r[4],
        "job_type": r[5],
        "asset_no": r[6],
        "problem": r[7],
        "solution": r[8],
        "complete_date": complete_date,
        "complete_time": complete_time,
        "close_note": r[10],
        "confirm_name": r[11],
    }

    return render_template("edit.html", data=data)


# ==================================================
# ลบงาน
# ==================================================
@app.route("/delete/<int:report_id>", methods=["POST"])
def delete_report(report_id):
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()

    # ดึงชื่อไฟล์ลายเซ็นก่อน (ถ้ามี)
    cursor.execute("SELECT signature FROM reports WHERE id=?", (report_id,))
    r = cursor.fetchone()

    # ลบข้อมูลใน DB
    cursor.execute("DELETE FROM reports WHERE id=?", (report_id,))

    conn.commit()
    conn.close()

    # ลบไฟล์ลายเซ็นออกจาก disk (ถ้ามี)
    if r and r[0]:
        try:
            sig_path = os.path.join("static", "signatures", r[0])
            if os.path.exists(sig_path):
                os.remove(sig_path)
        except:
            pass

    return redirect("/list?success=delete")


# ==================================================
# คัดลอกงาน
# ==================================================
@app.route("/copy/<int:report_id>", methods=["POST"])
def copy_report(report_id):
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()

    # ดึงข้อมูลต้นฉบับ (ไม่ใช้ SELECT *)
    cursor.execute(
        """
        SELECT
            receive_datetime,
            department,
            reporter,
            job_type,
            asset_no,
            problem,
            solution,
            completed_datetime,
            close_note,
            confirm_name,
            signature
        FROM reports
        WHERE id=?
    """,
        (report_id,),
    )

    r = cursor.fetchone()

    if not r:
        conn.close()
        return "ไม่พบข้อมูลงานต้นฉบับ", 404

    # สร้างเลขงานใหม่
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    work_no = generate_work_no(created_at)

    # INSERT งานใหม่ (ลายเซ็นไม่คัดลอก)
    cursor.execute(
        """
        INSERT INTO reports (
            work_no,
            receive_datetime,
            department,
            reporter,
            job_type,
            asset_no,
            problem,
            solution,
            completed_datetime,
            close_note,
            confirm_name,
            signature,
            created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """,
        (
            work_no,
            r[0],
            r[1],
            r[2],
            r[3],
            r[4],
            r[5],
            r[6],
            r[7],
            r[8],
            r[9],
            None,  # ❌ ไม่คัดลอกลายเซ็น
            created_at,
        ),
    )

    conn.commit()
    conn.close()

    return redirect("/list")

DEPT_FULLNAME = {
    "กกม.": "กลุ่มกฏหมาย",
    "กพร.": "กลุ่มพัฒนาระบบบริหาร",
    "กสส.": "กองส่งเสริมสิทธิและเสรีภาพ",
    "กยผ.": "กองยุทธศาสตร์และแผนสิทธิมนุษยชนแห่งชาติ",
    "สปท.": "สำนักงานป้องกันและปราบปรามการทรมานและการกระทำให้บุคคลสูญหาย",
    "กตน.": "กลุ่มตรวจสอบภายใน",
    "กสป.": "กองสิทธิมนุษยชนระหว่างประเทศ",
    "สชง.": "สำนักงานช่วยเหลือทางการเงินแก่ผู้เสียหายและจำเลยในคดีอาญา",
    "กสร.": "กองส่งเสริมการระงับข้อพิพาท",
    "กพส.": "กองพิทักษ์สิทธิและเสรีภาพ",
    "สคพ.": "สำนักงานคุ้มครองพยาน",
    "กชช.": "กลุ่มผู้เชี่ยวชาญ",
    "ศพท.": "ศูนย์ข้อมูลและพัฒนาระบบเทคโนโลยีสารสนเทศ",
    "กงค.": "กลุ่มงานคลัง",
    "กบค.": "กลุ่มงานบริหารและพัฒนาทรัพยากรบุคคล",
    "กชอ.": "กลุ่มงานช่วยอำนวยการและสื่อสารองค์กร",
    "กบพ.": "กลุ่มงานบริหารพัสดุ",
    "ห้องประชุม": "ห้องประชุม",
    "ห้องอธิบดี": "ห้องอธิบดี",
    "ห้องรองอธิบดี": "ห้องรองอธิบดี",
}

@app.route("/assets")
def assets_list():
    dept = request.args.get("dept")
    status = request.args.get("status")

    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()

    where = "WHERE 1=1"
    params = []

    # ===== filter หน่วยงาน =====
    if dept:
        where += " AND department = ?"
        params.append(dept)
        page_title = DEPT_FULLNAME.get(dept, dept)
    else:
        page_title = "รายการครุภัณฑ์ทั้งหมด"

    # ===== STEP 3: filter สถานะ =====
    if status:
        where += " AND status = ?"
        params.append(status)

    cursor.execute(f"""
        SELECT
            id,
            asset_no,
            asset_type,
            asset_model,
            serial_no,
            hostname,
            owner_name,
            department,
            status
        FROM assets
        {where}
        ORDER BY
            CASE asset_type
                WHEN 'Computer' THEN 1
                WHEN 'Notebook' THEN 2
                WHEN 'Printer' THEN 3
                WHEN 'Scanner' THEN 4
                WHEN 'Tablet' THEN 5
                WHEN 'UPS' THEN 6
                WHEN 'จอประชาสัมพันธ์' THEN 7
                ELSE 99
            END,
            LENGTH(
                REPLACE(
                    SUBSTR(asset_no, INSTR(asset_no, '/') + 1),
                    'O', ''
                )
            ),
            LENGTH(
                REPLACE(
                    SUBSTR(asset_no, 1, INSTR(asset_no, '/') - 1),
                    'I', ''
                )
            )
    """, params)



    assets = cursor.fetchall()
    conn.close()

    return render_template(
        "assets_list.html",
        assets=assets,
        page_title=page_title,
        current_dept=dept,
        status=status,              # ⭐ ส่งกลับไปให้ select จำค่า
        dept_map=DEPT_FULLNAME
    )


    
@app.route("/assets/add", methods=["GET", "POST"])
def add_asset():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()

    if request.method == "POST":
        try:
            cursor.execute("""
                INSERT INTO assets (
                    asset_no,
                    asset_type,
                    asset_model,
                    serial_no,
                    mac_address,
                    hostname,
                    owner_name,
                    position,
                    department,
                    status
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                request.form["asset_no"].strip(),
                request.form["asset_type"],
                request.form.get("asset_model"),
                request.form.get("serial_no"),
                request.form.get("mac_address"),
                request.form.get("hostname"),
                request.form.get("owner_name"),
                request.form.get("position"),
                request.form["department"],
                request.form["status"]
            ))

            conn.commit()
            conn.close()
            return redirect("/assets")

        except Exception as e:
            conn.close()
            return f"เกิดข้อผิดพลาด: {e}"

    conn.close()
    return render_template(
        "assets_add.html",
        dept_map=DEPT_FULLNAME
    )
   
    
@app.route("/assets/delete/<int:asset_id>")
def delete_asset(asset_id):
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM assets WHERE id = ?", (asset_id,))
    conn.commit()
    conn.close()
    return redirect("/assets?success=delete")

@app.route("/assets/edit/<int:asset_id>", methods=["GET", "POST"])
def edit_asset(asset_id):
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()

    if request.method == "POST":
        cursor.execute("""
            UPDATE assets SET
                asset_no = ?,
                asset_type = ?,
                asset_model = ?,
                serial_no = ?,
                mac_address = ?,
                hostname = ?,
                owner_name = ?,
                position = ?,
                department = ?,
                status = ?
            WHERE id = ?
        """, (
            request.form["asset_no"],
            request.form["asset_type"],
            request.form["asset_model"],
            request.form["serial_no"],
            request.form["mac_address"],
            request.form["hostname"],
            request.form["owner_name"],
            request.form["position"],
            request.form["department"],
            request.form["status"],
            asset_id
        ))
        conn.commit()
        conn.close()
        return redirect("/assets?success=edit")

    cursor.execute("SELECT * FROM assets WHERE id = ?", (asset_id,))
    asset = cursor.fetchone()
    conn.close()

    if not asset:
        return "ไม่พบข้อมูล", 404

    return render_template("assets_edit.html", asset=asset)

@app.route("/assets/export-excel")
def export_assets_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import send_file, request
    from datetime import datetime
    import io
    
    status = request.args.get("status")
    dept = request.args.get("dept")

        # ======================
    # ดึงข้อมูล (รองรับ dept + status)
    # ======================
    conn = get_db()
    cursor = conn.cursor()

    where = "WHERE 1=1"
    params = []

    if dept:
        where += " AND department = ?"
        params.append(dept)
        dept_full = DEPT_FULLNAME.get(dept, dept)
    else:
        dept_full = "รายการครุภัณฑ์ทั้งหมด"

    # ⭐ STEP 3: filter ตามสถานะ
    if status:
        where += " AND status = ?"
        params.append(status)

    cursor.execute(f"""
        SELECT
            asset_no,
            asset_type,
            asset_model,
            serial_no,
            hostname,
            owner_name,
            department,
            position
        FROM assets
        {where}
        ORDER BY
            CASE asset_type
                WHEN 'Computer' THEN 1
                WHEN 'Notebook' THEN 2
                WHEN 'Printer' THEN 3
                WHEN 'Scanner' THEN 4
                WHEN 'Tablet' THEN 5
                WHEN 'UPS' THEN 6
                WHEN 'จอประชาสัมพันธ์' THEN 7
                ELSE 99
            END,
            -- OO (ตัว O หลัง /) น้อย → มาก
            LENGTH(
                REPLACE(
                    SUBSTR(asset_no, INSTR(asset_no, '/') + 1),
                    'O', ''
                )
            ),
            -- I (ตัว I ก่อน /) น้อย → มาก
            LENGTH(
                REPLACE(
                    SUBSTR(asset_no, 1, INSTR(asset_no, '/') - 1),
                    'I', ''
                )
            )

    """, params)

    rows = cursor.fetchall()
    conn.close()

    # ======================
    # Excel
    # ======================
    wb = Workbook()
    ws = wb.active
    ws.title = "รายการครุภัณฑ์"

    # A4 แนวนอน (ไม่ fit)
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.print_title_rows = "3:3"

    # ===== style =====
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    font_title = Font(bold=True, size=14)
    font_header = Font(bold=True, size=10)
    font_normal = Font(size=9)

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    no_border = Border()

    # ======================
    # แถว 1 : ชื่อหน่วยงาน (ไม่มีขอบ)
    # ======================
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = dept_full
    c.font = font_title
    c.alignment = center
    ws.row_dimensions[1].height = 28

    # ======================
    # แถว 2 : เว้น (ไม่มีขอบ)
    # ======================
    ws.append([])
    ws.row_dimensions[2].height = 16

    # ======================
    # แถว 3 : หัวตาราง
    # ======================
    headers = [
        "ลำดับ",
        "เลขครุภัณฑ์",
        "ประเภท",
        "รุ่น",
        "Serial",
        "ชื่อเครื่อง",
        "ผู้ครอบครอง",
        "หน่วยงาน",
        "ตำแหน่ง",
    ]
    ws.append(headers)

    for col in range(1, 10):
        cell = ws.cell(row=3, column=col)
        cell.font = font_header
        cell.alignment = center
        cell.border = border

    ws.row_dimensions[3].height = 24

    # ======================
    # ข้อมูล
    # ======================
    start_row = 4
    for idx, r in enumerate(rows):
        row_no = start_row + idx

        ws.append([
            idx + 1,
            r[0],
            r[1],
            r[2],
            r[3],
            r[4],
            r[5],
            r[6],
            r[7],
        ])

        for col in range(1, 10):
            cell = ws.cell(row=row_no, column=col)
            cell.font = font_normal
            cell.alignment = center
            cell.border = border

        ws.row_dimensions[row_no].height = 20

    # ======================
    # เว้น 2 แถว + ลายเซ็นผู้ตรวจสอบ
    # ======================
    sign_row = ws.max_row + 3
    ws.merge_cells(start_row=sign_row, start_column=1, end_row=sign_row, end_column=9)

    sign = ws.cell(row=sign_row, column=1)
    sign.value = "ลายเซ็นผู้ตรวจสอบ ............................................................."
    sign.font = Font(size=10)
    sign.alignment = Alignment(horizontal="right", vertical="center")

    ws.row_dimensions[sign_row].height = 26

    # ======================
    # ความกว้างคอลัมน์ (ตามรูปที่ส่ง)
    # ======================
    widths = [
        6,   # A ลำดับ
        20,  # B เลขครุภัณฑ์
        12,  # C ประเภท
        18,  # D รุ่น
        16,  # E Serial
        18,  # F ชื่อเครื่อง
        20,  # G ผู้ครอบครอง
        10,  # H หน่วยงาน
        14,  # I ตำแหน่ง
    ]

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ======================
    # ส่งไฟล์
    # ======================
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"assets_A4_landscape_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/assets/summary")
def assets_summary():
    q = request.args.get("q", "").strip()

    conn = get_db()
    cursor = conn.cursor()

    # =========================
    # 1) ดึงข้อมูลสำหรับ Summary / Graph
    # =========================
    if q:
        cursor.execute("""
            SELECT department, asset_type
            FROM assets
            WHERE asset_no LIKE ?
        """, (f"%{q}%",))
    else:
        cursor.execute("""
            SELECT department, asset_type
            FROM assets
        """)

    rows = cursor.fetchall()
    
    
    # ===== mapping ชื่อย่อ -> ชื่อเต็ม =====
    dept_fullname = {}

    cursor.execute("""
        SELECT short_name, full_name
        FROM departments
        WHERE active = 1
    """)

    for short, full in cursor.fetchall():
        dept_fullname[short] = full
        
        

    # =========================
    # 2) สรุปข้อมูลตามหน่วยงาน
    # =========================
    summary = {}
    types = set()
    type_total = {}
    dept_total = {}

    for dept, a_type in rows:
        if dept not in summary:
            summary[dept] = {
                "items": {},
                "total": 0
            }

        summary[dept]["items"][a_type] = summary[dept]["items"].get(a_type, 0) + 1
        summary[dept]["total"] += 1

        types.add(a_type)
        type_total[a_type] = type_total.get(a_type, 0) + 1

    for dept, data in summary.items():
        dept_total[dept] = data["total"]

    types = sorted(types)

    # =========================
    # 3) Grand Total ทั้งตาราง
    # =========================
    grand_total = {
        "items": {},
        "total": 0
    }

    for dept, data in summary.items():
        for t, count in data["items"].items():
            grand_total["items"][t] = grand_total["items"].get(t, 0) + count
            grand_total["total"] += count

    # =========================
    # 4) 🔍 ค้นหา “เครื่องรายตัว” ตามเลขครุภัณฑ์
    # =========================
    asset_detail = None

    if q:
        cursor.execute("""
        SELECT
            asset_no,
            asset_type,
            asset_model,
            serial_no,
            hostname,
            owner_name,
            department,
            status
        FROM assets
        WHERE asset_no = ?
    """, (q,))
    asset_detail = cursor.fetchone()

    conn.close()

    # =========================
    # 5) ส่งไปที่ Template
    # =========================
    return render_template(
        "assets_summary.html",
        summary=summary,
        types=types,
        type_total=type_total,
        dept_total=dept_total,
        dept_detail=summary,
        grand_total=grand_total,
        asset_detail=asset_detail,   # ✅ เพิ่มตรงนี้
        dept_fullname=dept_fullname,   # 👈 เพิ่มบรรทัดนี้
        keyword=q
    )

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from flask import send_file
import io
import calendar
from collections import defaultdict

@app.route("/report-monthly-summary/export")
def export_report_monthly_summary():

    month = request.args.get("month", type=int)
    year = request.args.get("year", type=int)
    department_filter = request.args.get("department", "")

    last_day = calendar.monthrange(year, month)[1]
    date_from = f"{year}-{month:02d}-01"
    date_to = f"{year}-{month:02d}-{last_day}"

    # ================= DB =================
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    sql = """
        SELECT department, job_type, COUNT(*) AS total
        FROM reports
        WHERE receive_datetime >= ?
          AND receive_datetime <= ?
    """
    params = [f"{date_from} 00:00", f"{date_to} 23:59"]

    if department_filter:
        sql += " AND department = ?"
        params.append(department_filter)

    sql += " GROUP BY department, job_type"

    cur.execute(sql, params)
    rows = cur.fetchall()
    conn.close()

    # ================= เตรียมข้อมูล =================
    types = ["Software", "Hardware", "Network", "Other"]
    summary = defaultdict(lambda: {t: 0 for t in types})

    for r in rows:
        dept = (r["department"] or "ไม่ระบุ").strip()
        t = (r["job_type"] or "Other").strip().title()
        if t not in types:
            t = "Other"
        summary[dept][t] += r["total"]

    # ================= Excel =================
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Summary"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill(
        start_color="DBEAFE",  # ฟ้าอ่อน
        end_color="DBEAFE",
        fill_type="solid"
    )

    # ---------- หัวรายงาน ----------
    ws.merge_cells("A1:G1")
    ws["A1"] = "ตารางสรุปงานรายเดือน"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:G2")
    ws["A2"] = f"ประจำเดือน {thai_month_year(month, year)}"
    ws["A2"].alignment = center
    ws.row_dimensions[2].height = 26

    # ---------- หัวตาราง ----------
    headers = ["ลำดับ", "หน่วยงาน"] + types + ["รวม"]
    ws.append(headers)

    for col in range(1, 8):
        cell = ws.cell(row=3, column=col)
        cell.font = bold
        cell.alignment = center
        cell.border = border
        cell.fill = header_fill

    ws.row_dimensions[3].height = 30

    # ---------- เนื้อข้อมูล ----------
    row_idx = 4
    index = 1
    grand_total = {t: 0 for t in types}

    for dept, data in summary.items():
        total = sum(data.values())
        ws.append([index, dept] + [data[t] for t in types] + [total])

        for col in range(1, 8):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border
            cell.alignment = center

        ws.row_dimensions[row_idx].height = 26

        for t in types:
            grand_total[t] += data[t]

        index += 1
        row_idx += 1

    # ---------- แถวรวม ----------
    ws.append(
        ["", "รวมทั้งหมด"] +
        [grand_total[t] for t in types] +
        [sum(grand_total.values())]
    )

    for col in range(1, 8):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = bold
        cell.border = border
        cell.alignment = center

    ws.row_dimensions[row_idx].height = 30

    # ---------- ความกว้างคอลัมน์ ----------
    widths = [8, 24, 14, 14, 14, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w

    # ================= ส่งไฟล์ =================
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"ตารางสรุปงานรายเดือน_{thai_month_year(month, year)}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ==================================================
# เวลาเข้างาน
# ==================================================
@app.route("/attendance/checkin", methods=["POST"])
def attendance_checkin():
    staff_name = request.form.get("staff_name")

    if not staff_name:
        return redirect("/attendance")

    now = datetime.now(ZoneInfo("Asia/Bangkok"))

    work_date = now.strftime("%Y-%m-%d")
    time_now = now.strftime("%H:%M")

    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()

    cur.execute("""
        SELECT id FROM attendance
        WHERE staff_name = ? AND work_date = ?
    """, (staff_name, work_date))

    row = cur.fetchone()

    if row:
        cur.execute("""
            UPDATE attendance
            SET time_in = ?
            WHERE id = ?
        """, (time_now, row[0]))
    else:
        cur.execute("""
            INSERT INTO attendance (staff_name, work_date, time_in)
            VALUES (?, ?, ?)
        """, (staff_name, work_date, time_now))

    conn.commit()
    conn.close()

    return redirect(f"/attendance?staff_name={staff_name}")






# ==================================================
# เวลาออกงาน
# ==================================================
@app.route("/attendance/checkout", methods=["POST"])
def attendance_checkout():
    staff_name = request.form.get("staff_name")

    if not staff_name:
        return redirect("/attendance")

    now = datetime.now(ZoneInfo("Asia/Bangkok"))

    work_date = now.strftime("%Y-%m-%d")
    time_now = now.strftime("%H:%M")

    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()

    cur.execute("""
        UPDATE attendance
        SET time_out = ?
        WHERE staff_name = ? AND work_date = ?
    """, (time_now, staff_name, work_date))

    conn.commit()
    conn.close()

    return redirect(f"/attendance?staff_name={staff_name}")



# ==================================================
# เช็คงานบุคคล
# ==================================================
@app.route("/work-compare", methods=["GET"])
def work_compare():
    from datetime import datetime
    import sqlite3

    person1 = request.args.get("person1")
    person2 = request.args.get("person2")

    today = datetime.now()
    date_from = request.args.get(
        "date_from",
        f"{today.year}-{today.month:02d}-01"
    )
    date_to = request.args.get(
        "date_to",
        f"{today.year}-{today.month:02d}-31"
    )

    labels = ["Software", "Hardware", "Network", "Other"]

    # ✅ สร้าง data ก่อน
    data = {
        "labels": labels,
        "person1": [0, 0, 0, 0],
        "person2": [0, 0, 0, 0]
    }

    if person1 and person2:
        conn = sqlite3.connect(DB_NAME)
        cur = conn.cursor()

        cur.execute("""
            SELECT confirm_name, job_type, COUNT(*)
            FROM reports
            WHERE confirm_name IN (?, ?)
              AND receive_datetime BETWEEN ? AND ?
            GROUP BY confirm_name, job_type
        """, (
            person1, person2,
            f"{date_from} 00:00",
            f"{date_to} 23:59"
        ))

        rows = cur.fetchall()
        conn.close()

        index = {k: i for i, k in enumerate(labels)}

        for name, job, total in rows:
            if job not in index:
                continue
            if name == person1:
                data["person1"][index[job]] = total
            elif name == person2:
                data["person2"][index[job]] = total

    # ✅ คำนวณ “หลังจาก data พร้อมแล้ว”
    total1 = sum(data["person1"])
    total2 = sum(data["person2"])

    diff = [
        data["person1"][i] - data["person2"][i]
        for i in range(len(labels))
    ]
    diff_total = sum(diff)

    # ✅ return ต้องอยู่นอก if
    return render_template(
        "work_compare.html",
        data=data,
        person1=person1,
        person2=person2,
        date_from=date_from,
        date_to=date_to,
        total1=total1,
        total2=total2,
        diff=diff,
        diff_total=diff_total
    )

# ==================================================
# export ตารางครุภัณฑ์
# ==================================================
@app.route("/assets/export-summary")
def export_assets_summary():
    import sqlite3
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from flask import send_file
    from io import BytesIO
    from datetime import datetime

    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()

    cur.execute("""
        SELECT
            department,
            SUM(CASE WHEN asset_type='Computer' THEN 1 ELSE 0 END) AS Computer,
            SUM(CASE WHEN asset_type='Notebook' THEN 1 ELSE 0 END) AS Notebook,
            SUM(CASE WHEN asset_type='Printer' THEN 1 ELSE 0 END) AS Printer,
            SUM(CASE WHEN asset_type='Scanner' THEN 1 ELSE 0 END) AS Scanner,
            SUM(CASE WHEN asset_type='Tablet' THEN 1 ELSE 0 END) AS Tablet,
            SUM(CASE WHEN asset_type='UPS' THEN 1 ELSE 0 END) AS UPS,
            SUM(CASE WHEN asset_type='Display' THEN 1 ELSE 0 END) AS Display,
            COUNT(*) AS total
        FROM assets
        GROUP BY department
        ORDER BY department
    """)
    rows = cur.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "ตารางสรุปครุภัณฑ์"

    # ===== ตั้งค่าหน้ากระดาษ A4 แนวตั้ง =====
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.print_title_rows = "1:1"

    headers = [
        "หน่วยงาน", "Computer", "Notebook", "Printer",
        "Scanner", "Tablet", "UPS", "จอประชาสัมพันธ์", "รวม"
    ]
    ws.append(headers)

    # ===== style =====
    header_font = Font(bold=True, size=11)
    body_font = Font(size=10)
    bold_font = Font(bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # header
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # data
    for r in rows:
        ws.append(r)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = body_font
            cell.alignment = center
            cell.border = border

    # ===== แถวรวมล่างสุด =====
    ws.append([
        "รวมทั้งหมด",
        sum(r[1] for r in rows),
        sum(r[2] for r in rows),
        sum(r[3] for r in rows),
        sum(r[4] for r in rows),
        sum(r[5] for r in rows),
        sum(r[6] for r in rows),
        sum(r[7] for r in rows),
        sum(r[8] for r in rows),
    ])

    total_row = ws.max_row
    for cell in ws[total_row]:
        cell.font = bold_font
        cell.alignment = center
        cell.border = border

    # ===== ความสูงแถว (เล็กลง) =====
    for i in range(1, ws.max_row + 1):
        ws.row_dimensions[i].height = 20

    # ===== ความกว้างคอลัมน์ (คุมให้พอดี A4 แนวตั้ง) =====
    ws.column_dimensions["A"].width = 12   # หน่วยงาน (แคบ)
    ws.column_dimensions["B"].width = 9
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 9
    ws.column_dimensions["E"].width = 9
    ws.column_dimensions["F"].width = 9
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 12   # จอประชาสัมพันธ์ (แคบ)
    ws.column_dimensions["I"].width = 9    # รวม

    # ===== ส่งไฟล์ =====
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"asset_summary_A4_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ==================================================
# ใส่รหัส
# ==================================================
@app.route("/unlock", methods=["GET", "POST"])
def unlock():
    if request.method == "POST":
        password = request.form.get("password")

        if password == "123654":
            session["unlocked"] = True
            return redirect("/")
        else:
            return render_template(
                "unlock.html",
                error="รหัสไม่ถูกต้อง"
            )

    return render_template("unlock.html")


# ==================================================
# ชั่วคราว
# ==================================================



# ==================================================
# reset สถานะ
# ==================================================
@app.route("/assets/reset-check-status", methods=["POST"])
def reset_check_status():
    import sqlite3
    from flask import redirect, url_for, flash

    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()

    # รีเซ็ทเฉพาะครุภัณฑ์ที่ยังใช้งานอยู่
    cur.execute("""
        UPDATE assets
        SET status = 'ยังไม่ได้ตรวจสอบ'
        WHERE status = 'ใช้งาน'
    """)

    conn.commit()
    conn.close()

    flash("เริ่มรอบการตรวจสอบใหม่เรียบร้อยแล้ว", "success")
    return redirect(url_for("assets_list"))



# ==================================================
# RUN
# ==================================================
if __name__ == "__main__":
    init_db()              # สร้าง/เช็กตารางเดิม
    init_assets_db()       # สร้าง/เช็กตารางครุภัณฑ์
    fix_signature_column() # แก้โครงสร้างคอลัมน์เก่า
    app.run(host="0.0.0.0", debug=True)





